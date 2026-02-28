"""
Employee Daily Work Update Telegram Bot
========================================
Parses employee messages in format: EMP_ID Work description
Saves to Excel (local) + Google Sheets, sends notifications to Owner & HR.

Uses python-telegram-bot v21 API (ApplicationBuilder).
"""

import re
import os
import asyncio
import logging
from datetime import datetime

import pytz
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters

import config

# ─── Logging ─────────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─── Regex Pattern ───────────────────────────────────────────────────────────
# Matches: EMP_ID Work description
# EMP_ID: alphanumeric (e.g., PK01, DEV01)
# Work: everything after the first space
UPDATE_PATTERN = re.compile(
    r"^([A-Za-z0-9]+)\s+(.+)$",
    re.DOTALL,
)


# ─── Excel Functions ─────────────────────────────────────────────────────────
def save_to_excel(data: dict) -> bool:
    """Save the update to a professionally formatted Excel file."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        headers = [
            "Sr No", "Employee ID", "Department", "Employee Name",
            "Telegram Username", "Date", "Day", "Time",
            "Work Update", "Group Name",
        ]

        file_path = config.EXCEL_FILE

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            next_row = ws.max_row + 1
            sr_no = next_row - 1  # subtract header row
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Employee Updates"

            # ── Header styling ──
            header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="B0B0B0"),
                right=Side(style="thin", color="B0B0B0"),
                top=Side(style="thin", color="B0B0B0"),
                bottom=Side(style="thin", color="B0B0B0"),
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Column widths
            col_widths = [8, 15, 15, 20, 20, 14, 12, 10, 50, 25]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-filter
            ws.auto_filter.ref = f"A1:J1"

            next_row = 2
            sr_no = 1

        # ── Write data row ──
        row_data = [
            sr_no,
            data["emp_id"],
            data["department"],
            data["emp_name"],
            data["username"],
            data["date"],
            data["day"],
            data["time"],
            data["work_update"],
            data["group_name"],
        ]

        # Alternating row colors
        if sr_no % 2 == 0:
            row_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        data_font = Font(name="Arial", size=10)
        data_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Center-align Sr No, Date, Day, Time columns
        for col in [1, 6, 7, 8]:
            ws.cell(row=next_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        wb.save(file_path)
        logger.info(f"Excel: Saved update #{sr_no} to {file_path}")
        return True

    except Exception as e:
        logger.error(f"Excel save failed: {e}", exc_info=True)
        return False


# ─── Google Sheets Functions ─────────────────────────────────────────────────
def _save_to_google_sheets_sync(data: dict) -> bool:
    """Synchronous Google Sheets save (runs in a thread)."""
    if not config.GOOGLE_SHEET_ID:
        return False

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        creds = Credentials.from_service_account_file(config.GOOGLE_CREDS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(config.GOOGLE_SHEET_ID).sheet1

        # Check if headers exist
        existing = sheet.get_all_values()
        headers = [
            "Sr No", "Employee ID", "Department", "Employee Name",
            "Telegram Username", "Date", "Day", "Time",
            "Work Update", "Group Name",
        ]

        if not existing:
            sheet.append_row(headers)
            sr_no = 1
        else:
            sr_no = len(existing)  # rows minus header = count, +1 for new = len

        row = [
            sr_no,
            data["emp_id"],
            data["department"],
            data["emp_name"],
            data["username"],
            data["date"],
            data["day"],
            data["time"],
            data["work_update"],
            data["group_name"],
        ]

        sheet.append_row(row)
        logger.info(f"Google Sheets: Saved update #{sr_no}")
        return True

    except Exception as e:
        logger.error(f"Google Sheets save failed: {e}", exc_info=True)
        return False


async def save_to_google_sheets(data: dict) -> bool:
    """Save to Google Sheets in a background thread (non-blocking)."""
    try:
        return await asyncio.to_thread(_save_to_google_sheets_sync, data)
    except Exception as e:
        logger.error(f"Google Sheets async wrapper failed: {e}", exc_info=True)
        return False


# ─── Notification Functions ──────────────────────────────────────────────────
async def send_personal_notification(bot, chat_id: str, data: dict, label: str):
    """Send a personal notification to Owner or HR."""
    if not chat_id:
        return

    try:
        message = (
            f"📬 *New Work Update Received*\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"👤 *Employee:* {data['emp_name']} (`{data['emp_id']}`)\n"
            f"🏢 *Department:* {data['department']}\n"
            f"💬 *Telegram:* @{data['username']}\n"
            f"📅 *Date:* {data['date']} ({data['day']})\n"
            f"🕐 *Time:* {data['time']}\n"
            f"📝 *Update:*\n{data['work_update']}\n"
            f"━━━━━━━━━━━━━━━━━━━━━━\n"
            f"📍 *Group:* {data['group_name']}"
        )
        await bot.send_message(
            chat_id=int(chat_id),
            text=message,
            parse_mode="Markdown",
        )
        logger.info(f"Notification sent to {label} (Chat ID: {chat_id})")

    except Exception as e:
        logger.warning(f"Failed to notify {label}: {e}")


# ─── Command Handlers ────────────────────────────────────────────────────────
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Welcome message for /start."""
    message = (
        "👋 *Welcome to the Employee Work Update Bot!*\n\n"
        "📝 *How to submit your daily update:*\n"
        "`YOUR_ID Your work description`\n\n"
        "📌 *Example:*\n"
        "`DEV01 Fixed the login page and tested it`\n\n"
        "💡 Use /help to see all commands."
    )
    await update.message.reply_text(message, parse_mode="Markdown")
    logger.info(f"Start command used by {update.effective_user.id}")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Display all available commands and usage."""
    user_id = str(update.effective_user.id)
    is_admin = user_id in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]

    message = (
        "📖 *Bot Help & Commands*\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "📝 *Submit Your Update:*\n"
        "`YOUR_ID Your work description`\n\n"
        "📌 *Example:*\n"
        "`DEV01 Completed API integration and unit tests`\n\n"
        "⚠️ *Rules:*\n"
        "• Your Employee ID must be registered\n"
        "• Only one submission per day\n"
        "• Contact admin if you need to re-submit\n\n"
        "🔧 *Commands:*\n"
        "• /start — Welcome message\n"
        "• /help — This help menu\n"
    )

    if is_admin:
        message += (
            "\n👑 *Admin Commands:*\n"
            "• /staff — View all registered employees\n"
            "• /addstaff `ID - Name - Dept` — Add employee\n"
            "• /removestaff `ID` — Remove employee\n"
            "• /allow `ID` — Allow re-submission today\n"
            "• /report — Today's submission status\n"
        )

    await update.message.reply_text(message, parse_mode="Markdown")
    logger.info(f"Help command used by {update.effective_user.id}")


async def staff_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Display the list of registered staff members and instructions (Owner/HR only)."""
    user_id = str(update.effective_user.id)
    
    # Check if user is Owner or HR
    if user_id not in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]:
        await update.message.reply_text("🚫 *Permission Denied:* Only the Owner and HR can use this command.", parse_mode="Markdown")
        logger.warning(f"Unauthorized attempt by {user_id} to access /staff")
        return

    if not config.STAFF_RECORDS:
        await update.message.reply_text("📋 No staff records configured.")
        return

    message = "👥 *Registered Staff & Attendance Format*\n"
    message += "━━━━━━━━━━━━━━━━━━━━━━\n"
    message += "Use the following ID for your daily updates:\n\n"
    
    for emp_id, info in config.STAFF_RECORDS.items():
        message += f"• `{emp_id}` — {info['name']} ({info['dept']})\n"
    
    message += "\n📝 *Update Format:*\n"
    message += "`ID Your work description`"
    
    await update.message.reply_text(message, parse_mode="Markdown")
    logger.info("Staff list shared")


async def addstaff_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Dynamically add a new staff member (Owner/HR only)."""
    user_id = str(update.effective_user.id)
    
    # Check if user is Owner or HR
    if user_id not in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]:
        await update.message.reply_text("🚫 *Permission Denied:* Only the Owner and HR can add staff.", parse_mode="Markdown")
        return

    # Expected: /addstaff ID - Name - Dept
    args = " ".join(context.args)
    if not args:
        await update.message.reply_text("❗ *Usage:* `/addstaff ID - Name - Department`", parse_mode="Markdown")
        return

    try:
        # Simple split by dash
        parts = [p.strip() for p in args.split("-")]
        if len(parts) != 3:
            raise ValueError("Invalid format")
        
        emp_id, name, dept = parts
        
        # Save to file
        success = config.save_staff_record(emp_id, name, dept)
        
        if success:
            # Refresh in-memory records
            config.STAFF_RECORDS = config.load_staff_records()
            await update.message.reply_text(
                f"✅ *Staff Added Successfully!*\n\n"
                f"🆔 *ID:* `{emp_id.upper()}`\n"
                f"👤 *Name:* {name}\n"
                f"🏢 *Dept:* {dept.upper()}",
                parse_mode="Markdown"
            )
            logger.info(f"Added new staff: {emp_id} - {name}")
        else:
            await update.message.reply_text("❌ *Error:* Failed to save staff record.", parse_mode="Markdown")

    except Exception:
        await update.message.reply_text("❗ *Format error:* Use `/addstaff ID - Name - Department`", parse_mode="Markdown")


async def removestaff_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Remove a staff member (Owner/HR only)."""
    user_id = str(update.effective_user.id)
    
    # Check if user is Owner or HR
    if user_id not in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]:
        await update.message.reply_text("🚫 *Permission Denied:* Only the Owner and HR can remove staff.", parse_mode="Markdown")
        return

    if not context.args:
        await update.message.reply_text("❗ *Usage:* `/removestaff EMP_ID`", parse_mode="Markdown")
        return

    emp_id = context.args[0].upper()
    
    # Check if employee exists
    if emp_id not in config.STAFF_RECORDS:
        await update.message.reply_text(f"❌ Employee `{emp_id}` not found in records.", parse_mode="Markdown")
        return

    staff_name = config.STAFF_RECORDS[emp_id]["name"]
    success = config.remove_staff_record(emp_id)
    
    if success:
        # Refresh in-memory records
        config.STAFF_RECORDS = config.load_staff_records()
        await update.message.reply_text(
            f"✅ *Staff Removed Successfully!*\n\n"
            f"🆔 *ID:* `{emp_id}`\n"
            f"👤 *Name:* {staff_name}",
            parse_mode="Markdown"
        )
        logger.info(f"Removed staff: {emp_id} - {staff_name}")
    else:
        await update.message.reply_text("❌ *Error:* Failed to remove staff record.", parse_mode="Markdown")


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show today's submission status (Owner/HR only)."""
    user_id = str(update.effective_user.id)
    
    # Check if user is Owner or HR
    if user_id not in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]:
        await update.message.reply_text("🚫 *Permission Denied:* Only the Owner and HR can view reports.", parse_mode="Markdown")
        return

    tz = pytz.timezone(config.TIMEZONE)
    now = datetime.now(tz)
    today_str = now.strftime("%Y-%m-%d")

    # Get daily log
    if "daily_log" not in context.bot_data:
        context.bot_data["daily_log"] = config.load_daily_log()

    today_log = context.bot_data.get("daily_log", {}).get(today_str, {})
    
    submitted = []
    not_submitted = []
    
    for emp_id, info in config.STAFF_RECORDS.items():
        if emp_id in today_log:
            submitted.append(f"✅ `{emp_id}` — {info['name']}")
        else:
            not_submitted.append(f"❌ `{emp_id}` — {info['name']}")

    total = len(config.STAFF_RECORDS)
    done = len(submitted)

    message = (
        f"📊 *Daily Report — {now.strftime('%d %b %Y')}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📈 *Progress:* {done}/{total} submitted\n\n"
    )

    if submitted:
        message += "*✅ Submitted:*\n"
        message += "\n".join(submitted)
        message += "\n\n"

    if not_submitted:
        message += "*❌ Not Submitted:*\n"
        message += "\n".join(not_submitted)

    if not submitted and not not_submitted:
        message += "No staff records found."

    await update.message.reply_text(message, parse_mode="Markdown")
    logger.info(f"Report generated: {done}/{total} submitted")


# ─── Main Message Handler ────────────────────────────────────────────────────
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Process incoming group messages and parse work updates."""
    if not update.message or not update.message.text:
        return

    text = update.message.text.strip()
    match = UPDATE_PATTERN.match(text)

    if not match:
        # Silently ignore messages that don't match the format
        return

    # ── Parse the message ──
    emp_id = match.group(1).strip().upper()
    work_update = match.group(2).strip()

    # ── FIX 1: Only process REGISTERED employee IDs ──
    # This prevents "Hello everyone" from being treated as a work update
    staff_info = config.STAFF_RECORDS.get(emp_id)
    if not staff_info:
        # Not a registered employee ID — silently ignore
        return

    emp_name = staff_info['name']
    department = staff_info['dept']

    # Timezone-aware datetime
    tz = pytz.timezone(config.TIMEZONE)
    now = datetime.now(tz)

    # Build data dict
    user = update.message.from_user
    username = user.username if user.username else user.first_name or "Unknown"

    group_name = ""
    if update.message.chat.title:
        group_name = update.message.chat.title
    elif update.message.chat.type == "private":
        group_name = "Private Chat"

    data = {
        "emp_id": emp_id,
        "department": department,
        "emp_name": emp_name,
        "username": username,
        "date": now.strftime("%d-%m-%Y"),
        "day": now.strftime("%A"),
        "time": now.strftime("%I:%M %p"),
        "work_update": work_update,
        "group_name": group_name,
    }

    logger.info(f"New update: {emp_id} | {department} | {emp_name} | {work_update[:80]}")

    # ── FIX 3: Load persistent daily log ──
    if "daily_log" not in context.bot_data:
        context.bot_data["daily_log"] = config.load_daily_log()

    today_str = now.strftime("%Y-%m-%d")
    if today_str not in context.bot_data["daily_log"]:
        context.bot_data["daily_log"][today_str] = {}

    if emp_id in context.bot_data["daily_log"][today_str]:
        # Already submitted today
        await update.message.reply_text(
            f"❌ *Already Submitted*\n\n{emp_name} (`{emp_id}`) has already submitted an update for today. "
            f"Please contact an admin if you need to re-submit.",
            parse_mode="Markdown"
        )
        logger.info(f"Duplicate submission blocked: {emp_id}")
        return

    # ── Save to Excel ──
    excel_saved = save_to_excel(data)

    # ── FIX 2: Save to Google Sheets (non-blocking) ──
    sheets_saved = await save_to_google_sheets(data)

    # ── FIX 4: Log warning if Excel saved on Railway ──
    if config.IS_RAILWAY and excel_saved:
        logger.warning("Excel saved on Railway — file will be lost on next deploy. Use Google Sheets for persistence.")

    # ── Build confirmation message ──
    confirmation = f"Thank you, *{data['emp_name']}*! ✅"

    # ── Send group confirmation ──
    try:
        await update.message.reply_text(confirmation, parse_mode="Markdown")
        logger.info("Group confirmation sent")
    except Exception as e:
        logger.warning(f"Failed to send group confirmation: {e}")

    # ── Send personal notifications ──
    await send_personal_notification(context.bot, config.OWNER_CHAT_ID, data, "Owner")
    await send_personal_notification(context.bot, config.HR_CHAT_ID, data, "HR")

    # ── Record Submission (in-memory + file) ──
    context.bot_data["daily_log"][today_str][emp_id] = True
    config.save_daily_log(context.bot_data["daily_log"])

    logger.info("Update processing complete")


async def allow_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Allow an employee to re-submit today."""
    user_id = str(update.effective_user.id)
    if user_id not in [str(config.OWNER_CHAT_ID), str(config.HR_CHAT_ID)]:
        await update.message.reply_text("🚫 *Permission Denied:* Only admins can use this command.", parse_mode="Markdown")
        return

    if not context.args:
        await update.message.reply_text("Usage: `/allow EMP_ID`", parse_mode="Markdown")
        return

    emp_id = context.args[0].upper()
    now = datetime.now(pytz.timezone(config.TIMEZONE))
    today_str = now.strftime("%Y-%m-%d")

    if "daily_log" not in context.bot_data:
        context.bot_data["daily_log"] = config.load_daily_log()

    if today_str in context.bot_data.get("daily_log", {}) and emp_id in context.bot_data["daily_log"][today_str]:
        del context.bot_data["daily_log"][today_str][emp_id]
        config.save_daily_log(context.bot_data["daily_log"])
        await update.message.reply_text(f"✅ Employee `{emp_id}` can now re-submit today.", parse_mode="Markdown")
        logger.info(f"Re-submission allowed for {emp_id}")
        return

    await update.message.reply_text(f"Employee `{emp_id}` has not submitted anything today.", parse_mode="Markdown")


# ─── Bot Startup ─────────────────────────────────────────────────────────────
def main():
    """Start the bot."""
    logger.info("=" * 55)
    logger.info("  Employee Work Update Bot — Starting")
    logger.info("=" * 55)

    # Validate config
    if not config.BOT_TOKEN:
        logger.error("BOT_TOKEN is not set! Set it in .env file or as environment variable.")
        return

    logger.info(f"  Excel file: {config.EXCEL_FILE}")
    logger.info(f"  Google Sheets: {'Configured' if config.GOOGLE_SHEET_ID else 'Disabled'}")
    logger.info(f"  Owner notifications: {'ON' if config.OWNER_CHAT_ID else 'OFF'}")
    logger.info(f"  HR notifications: {'ON' if config.HR_CHAT_ID else 'OFF'}")
    logger.info(f"  Timezone: {config.TIMEZONE}")
    logger.info(f"  Staff records: {len(config.STAFF_RECORDS)} employees loaded")

    if config.IS_RAILWAY:
        logger.warning("Running on Railway — Excel files are ephemeral. Use Google Sheets for persistent storage.")

    logger.info("=" * 55)
    logger.info("  Bot is running... Press Ctrl+C to stop")

    # Build and run the application (v21 API)
    app = ApplicationBuilder().token(config.BOT_TOKEN).build()

    # Add command handlers
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("staff", staff_command))
    app.add_handler(CommandHandler("addstaff", addstaff_command))
    app.add_handler(CommandHandler("removestaff", removestaff_command))
    app.add_handler(CommandHandler("allow", allow_command))
    app.add_handler(CommandHandler("report", report_command))

    # Handle all text messages (groups + private)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Start polling
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
