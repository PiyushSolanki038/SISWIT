"""
Excel & Google Sheets handler for the Employee Work Update Bot.
Handles saving, formatting, dashboard, leave register, and monthly sheets.
"""

import os
import asyncio
import logging
from datetime import datetime

import config

logger = logging.getLogger(__name__)

# ─── Excel Constants ─────────────────────────────────────────────────────────
HEADERS = [
    "Sr No", "Employee ID", "Department", "Employee Name",
    "Telegram Username", "Date", "Day", "Time",
    "Work Update", "Group Name", "Status", "On Time",
]


def _get_month_sheet_name(date_obj=None):
    """Get the sheet name for the current month, e.g. 'Mar 2026'."""
    if date_obj is None:
        import pytz
        tz = pytz.timezone(config.TIMEZONE)
        date_obj = datetime.now(tz)
    return date_obj.strftime("%b %Y")


def _is_on_time(time_str):
    """Check if submission time is before the deadline."""
    try:
        deadline = config.get_deadline()
        deadline_h, deadline_m = map(int, deadline.split(":"))
        # Parse time like "09:30 AM"
        sub_time = datetime.strptime(time_str, "%I:%M %p")
        deadline_time = sub_time.replace(hour=deadline_h, minute=deadline_m)
        return sub_time <= deadline_time
    except Exception:
        return True  # Default to on-time if parsing fails


def save_to_excel(data: dict) -> bool:
    """Save the update to a professionally formatted Excel file with monthly sheets."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        file_path = config.EXCEL_FILE
        month_name = _get_month_sheet_name()
        on_time = _is_on_time(data["time"])

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Get or create monthly sheet
        if month_name in wb.sheetnames:
            ws = wb[month_name]
            next_row = ws.max_row + 1
            sr_no = next_row - 1
        else:
            ws = wb.create_sheet(month_name, 0)
            _format_header(ws)
            next_row = 2
            sr_no = 1

        # Write data row
        row_data = [
            sr_no, data["emp_id"], data["department"], data["emp_name"],
            data["username"], data["date"], data["day"], data["time"],
            data["work_update"], data["group_name"],
            "✅ Present", "✅ Yes" if on_time else "❌ Late",
        ]

        # Styling
        if sr_no % 2 == 0:
            row_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        data_font = Font(name="Arial", size=10)
        data_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Center-align specific columns
        for col in [1, 6, 7, 8, 11, 12]:
            ws.cell(row=next_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Color the On Time cell
        on_time_cell = ws.cell(row=next_row, column=12)
        if on_time:
            on_time_cell.font = Font(name="Arial", size=10, color="228B22")
        else:
            on_time_cell.font = Font(name="Arial", size=10, color="DC143C")

        # Update dashboard
        _update_dashboard(wb)

        wb.save(file_path)
        logger.info(f"Excel: Saved update #{sr_no} to {month_name}")
        return True

    except Exception as e:
        logger.error(f"Excel save failed: {e}", exc_info=True)
        return False


def _format_header(ws):
    """Apply professional header formatting to a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    col_widths = [8, 15, 15, 20, 20, 14, 12, 10, 50, 25, 12, 12]
    for i, width in enumerate(col_widths, 1):
        col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(64 + (i - 1) % 26 + 1)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L1"


def _update_dashboard(wb):
    """Create or update the Dashboard summary sheet."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    dash_name = "Dashboard"
    if dash_name in wb.sheetnames:
        ws = wb[dash_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(dash_name)

    # Dashboard headers
    dash_headers = ["Employee ID", "Name", "Department", "Days Submitted", "Late Count", "On Leave", "Attendance %"]
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(dash_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Count submissions per employee across all monthly sheets
    emp_stats = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == dash_name:
            continue
        sheet = wb[sheet_name]
        for row in range(2, sheet.max_row + 1):
            emp_id = sheet.cell(row=row, column=2).value
            if not emp_id:
                continue
            if emp_id not in emp_stats:
                emp_stats[emp_id] = {"days": 0, "late": 0, "leave": 0}
            emp_stats[emp_id]["days"] += 1
            on_time_val = sheet.cell(row=row, column=12).value
            if on_time_val and "Late" in str(on_time_val):
                emp_stats[emp_id]["late"] += 1
            status_val = sheet.cell(row=row, column=11).value
            if status_val and "Leave" in str(status_val):
                emp_stats[emp_id]["leave"] += 1

    # Write rows
    row_idx = 2
    for emp_id, info in config.STAFF_RECORDS.items():
        stats = emp_stats.get(emp_id, {"days": 0, "late": 0, "leave": 0})
        # Simple 30-day attendance
        attendance = round((stats["days"] / 30) * 100, 1) if stats["days"] > 0 else 0

        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=info["name"])
        ws.cell(row=row_idx, column=3, value=info["dept"])
        ws.cell(row=row_idx, column=4, value=stats["days"])
        ws.cell(row=row_idx, column=5, value=stats["late"])
        ws.cell(row=row_idx, column=6, value=stats["leave"])
        ws.cell(row=row_idx, column=7, value=f"{attendance}%")

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_idx, column=col).font = Font(name="Arial", size=10)

        row_idx += 1

    # Column widths
    for i, w in enumerate([15, 20, 15, 16, 12, 12, 14], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"


def count_monthly_leaves(emp_id, leave_date_str):
    """Count how many approved leaves an employee has in the same month."""
    try:
        leave_date = datetime.strptime(leave_date_str, "%d-%m-%Y")
        month_key = leave_date.strftime("%Y-%m")
        leave_log = config.load_leave_log()
        count = 0
        for date_str, entries in leave_log.items():
            if date_str.startswith(month_key) and emp_id in entries:
                count += 1
        return count
    except Exception:
        return 0


def save_leave_to_excel(emp_id, emp_name, dept, leave_date, reason, approved_by, leave_count):
    """Save an approved leave entry to the Leave Register sheet in Excel.
    If leave_count > 3, adds -₹500 salary deduction per extra leave.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        file_path = config.EXCEL_FILE

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        sheet_name = "Leave Register"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            next_row = ws.max_row + 1

            # Auto-upgrade old sheets: add missing columns if they don't exist
            if ws.cell(row=1, column=9).value != "Leave #":
                header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
                header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
                for col_idx, h in [(9, "Leave #"), (10, "Deduction")]:
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions["I"].width = 10
                ws.column_dimensions["J"].width = 16
        else:
            ws = wb.create_sheet(sheet_name)
            leave_headers = [
                "Sr No", "Employee ID", "Name", "Department",
                "Leave Date", "Reason", "Approved By", "Status",
                "Leave #", "Deduction",
            ]
            header_fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
            header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

            for col_idx, h in enumerate(leave_headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, w in enumerate([8, 15, 20, 15, 14, 40, 20, 14, 10, 16], 1):
                ws.column_dimensions[chr(64 + i)].width = w

            ws.freeze_panes = "A2"
            next_row = 2

        # Calculate deduction
        is_extra = leave_count > 3
        deduction = f"-₹{(leave_count - 3) * 500}" if is_extra else "—"
        status = "⚠️ Extra Leave" if is_extra else "✅ Approved"

        sr = next_row - 1
        row_data = [
            sr, emp_id, emp_name, dept, leave_date,
            reason, approved_by, status, leave_count, deduction,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Highlight deduction in red if applicable
        if is_extra:
            deduction_cell = ws.cell(row=next_row, column=10)
            deduction_cell.font = Font(name="Arial", size=10, bold=True, color="DC143C")
            deduction_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

            status_cell = ws.cell(row=next_row, column=8)
            status_cell.font = Font(name="Arial", size=10, color="DC143C")

        wb.save(file_path)
        logger.info(f"Excel: Leave #{leave_count} recorded for {emp_id} on {leave_date}" +
                     (f" (Deduction: {deduction})" if is_extra else ""))
        return is_extra
    except Exception as e:
        logger.error(f"Excel leave save failed: {e}", exc_info=True)
        return False


# ─── Google Sheets Functions ─────────────────────────────────────────────────
def _save_to_google_sheets_sync(data: dict) -> bool:
    """Synchronous Google Sheets save (runs in a thread)."""
    if not config.GOOGLE_SHEET_ID:
        return False

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        creds = Credentials.from_service_account_file(config.GOOGLE_CREDS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(config.GOOGLE_SHEET_ID)

        month_name = _get_month_sheet_name()
        on_time = _is_on_time(data["time"])

        # Get or create monthly worksheet
        try:
            sheet = spreadsheet.worksheet(month_name)
        except gspread.exceptions.WorksheetNotFound:
            sheet = spreadsheet.add_worksheet(title=month_name, rows=1000, cols=12)
            sheet.append_row(HEADERS)
            # Format header
            sheet.format("A1:L1", {
                "backgroundColor": {"red": 0.106, "green": 0.227, "blue": 0.361},
                "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True, "fontSize": 11},
                "horizontalAlignment": "CENTER",
            })
            sheet.freeze(rows=1)

        existing = sheet.get_all_values()
        sr_no = len(existing) if existing else 1

        row = [
            sr_no, data["emp_id"], data["department"], data["emp_name"],
            data["username"], data["date"], data["day"], data["time"],
            data["work_update"], data["group_name"],
            "✅ Present", "✅ Yes" if on_time else "❌ Late",
        ]

        sheet.append_row(row)
        logger.info(f"Google Sheets: Saved update #{sr_no} to {month_name}")
        return True

    except Exception as e:
        logger.error(f"Google Sheets save failed: {e}", exc_info=True)
        return False


async def save_to_google_sheets(data: dict) -> bool:
    """Save to Google Sheets in a background thread (non-blocking)."""
    try:
        return await asyncio.to_thread(_save_to_google_sheets_sync, data)
    except Exception as e:
        logger.error(f"Google Sheets async wrapper failed: {e}", exc_info=True)
        return False


def _save_leave_to_google_sheets_sync(emp_id, emp_name, dept, leave_date, reason, approved_by, leave_count):
    """Save leave entry to Google Sheets Leave Register (synchronous)."""
    if not config.GOOGLE_SHEET_ID:
        return False

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]

        creds = Credentials.from_service_account_file(config.GOOGLE_CREDS_FILE, scopes=scopes)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(config.GOOGLE_SHEET_ID)

        sheet_name = "Leave Register"
        leave_headers = [
            "Sr No", "Employee ID", "Name", "Department",
            "Leave Date", "Reason", "Approved By", "Status",
            "Leave #", "Deduction",
        ]

        # Get or create Leave Register sheet
        try:
            sheet = spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            sheet = spreadsheet.add_worksheet(title=sheet_name, rows=500, cols=10)
            sheet.append_row(leave_headers)
            # Format header
            sheet.format("A1:J1", {
                "backgroundColor": {"red": 0.902, "green": 0.318, "blue": 0.0},
                "textFormat": {
                    "foregroundColor": {"red": 1, "green": 1, "blue": 1},
                    "bold": True, "fontSize": 11,
                },
                "horizontalAlignment": "CENTER",
            })
            sheet.freeze(rows=1)

        # Calculate deduction
        is_extra = leave_count > 3
        deduction = f"-₹{(leave_count - 3) * 500}" if is_extra else "—"
        status = "⚠️ Extra Leave" if is_extra else "✅ Approved"

        existing = sheet.get_all_values()
        sr_no = len(existing) if existing else 1

        row = [
            sr_no, emp_id, emp_name, dept, leave_date,
            reason, approved_by, status, leave_count, deduction,
        ]

        sheet.append_row(row)

        # Color the deduction cell red if extra leave
        if is_extra:
            row_num = len(sheet.get_all_values())
            # Color deduction cell red
            sheet.format(f"J{row_num}", {
                "textFormat": {
                    "foregroundColor": {"red": 0.86, "green": 0.08, "blue": 0.24},
                    "bold": True,
                },
                "backgroundColor": {"red": 1, "green": 0.88, "blue": 0.88},
            })
            # Color status cell red
            sheet.format(f"H{row_num}", {
                "textFormat": {
                    "foregroundColor": {"red": 0.86, "green": 0.08, "blue": 0.24},
                },
            })

        logger.info(f"Google Sheets: Leave #{leave_count} saved for {emp_id}")
        return True

    except Exception as e:
        logger.error(f"Google Sheets leave save failed: {e}", exc_info=True)
        return False


async def save_leave_to_google_sheets(emp_id, emp_name, dept, leave_date, reason, approved_by, leave_count):
    """Save leave to Google Sheets in a background thread (non-blocking)."""
    try:
        return await asyncio.to_thread(
            _save_leave_to_google_sheets_sync,
            emp_id, emp_name, dept, leave_date, reason, approved_by, leave_count
        )
    except Exception as e:
        logger.error(f"Google Sheets leave async failed: {e}", exc_info=True)
        return False
